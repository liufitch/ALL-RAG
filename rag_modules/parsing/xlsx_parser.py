from __future__ import annotations

import io
from collections.abc import Iterator, Mapping
from dataclasses import dataclass
from pathlib import Path
from pathlib import PurePosixPath
from typing import Any, BinaryIO
from urllib.parse import urlsplit
from zipfile import BadZipFile, ZipFile

from lxml import etree
from lxml.etree import XMLSyntaxError
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from rag_modules.config.settings import settings
from rag_modules.parsing.base import ParseContext
from rag_modules.parsing.models import DocumentParseError, ParsedDocument, ParserWarning
from rag_modules.parsing.tabular import TableRowBudget, table_blocks


_WORKBOOK_PART = "xl/workbook.xml"
_WORKBOOK_RELATIONSHIPS_PART = "xl/_rels/workbook.xml.rels"
_TRANSITIONAL_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_STRICT_RELATIONSHIP_NAMESPACE = (
    "http://purl.oclc.org/ooxml/officeDocument/relationships"
)
_WORKSHEET_RELATIONSHIP_TYPES = frozenset(
    {
        f"{_TRANSITIONAL_RELATIONSHIP_NAMESPACE}/worksheet",
        f"{_STRICT_RELATIONSHIP_NAMESPACE}/worksheet",
    }
)
_RELATIONSHIP_NAMESPACES = frozenset(
    {_TRANSITIONAL_RELATIONSHIP_NAMESPACE, _STRICT_RELATIONSHIP_NAMESPACE}
)
_XLSX_MAX_ROW = 1_048_576
_XLSX_MAX_COLUMN = 16_384


@dataclass(frozen=True)
class _WorksheetPreflight:
    title: str
    part_name: str
    physical_coordinates: frozenset[tuple[int, int]]


class XlsxParser:
    """Read visible XLSX sheets without evaluating formulas or expanding dimensions."""

    source_type = "xlsx"

    def parse(self, stream: BinaryIO, context: ParseContext) -> ParsedDocument:
        if Path(context.filename).suffix.lower() != ".xlsx":
            raise DocumentParseError(
                "UNSUPPORTED_FILE_TYPE", "XlsxParser only supports .xlsx files."
            )
        payload = stream.read()
        try:
            worksheet_preflights = _preflight_worksheet_xml(payload)
            formula_workbook = load_workbook(
                io.BytesIO(payload), read_only=False, data_only=False, keep_links=False
            )
            cached_workbook = load_workbook(
                io.BytesIO(payload), read_only=False, data_only=True, keep_links=False
            )
        except DocumentParseError:
            raise
        except (
            BadZipFile,
            InvalidFileException,
            OSError,
            KeyError,
            ValueError,
            XMLSyntaxError,
        ) as error:
            raise _malformed_xlsx_error() from error

        try:
            blocks = []
            warnings: list[ParserWarning] = []
            budget = TableRowBudget()
            formula_worksheets = formula_workbook.worksheets
            cached_worksheets = cached_workbook.worksheets
            if not (
                len(formula_worksheets)
                == len(cached_worksheets)
                == len(worksheet_preflights)
            ):
                raise _malformed_xlsx_error()
            worksheet_triples = tuple(
                zip(worksheet_preflights, formula_worksheets, cached_worksheets)
            )
            if any(
                preflight.title != worksheet.title
                or preflight.title != cached_worksheet.title
                for preflight, worksheet, cached_worksheet in worksheet_triples
            ):
                raise _malformed_xlsx_error()
            for preflight, worksheet, cached_worksheet in worksheet_triples:
                physical_cells = _physical_cells(
                    worksheet,
                    cached_worksheet,
                    preflight.physical_coordinates,
                )
                if worksheet.sheet_state != "visible":
                    for _physical_cell in physical_cells:
                        pass
                    warnings.append(_hidden_sheet_warning(worksheet.title))
                    continue
                sheet_blocks = table_blocks(
                    _worksheet_rows(
                        worksheet,
                        physical_cells,
                        warnings,
                    ),
                    worksheet.title,
                    budget,
                )
                if not sheet_blocks:
                    warnings.append(_empty_sheet_warning(worksheet.title))
                blocks.extend(sheet_blocks)
        except DocumentParseError:
            raise
        except (BadZipFile, OSError, KeyError, ValueError, XMLSyntaxError) as error:
            raise _malformed_xlsx_error() from error
        finally:
            formula_workbook.close()
            cached_workbook.close()

        if not blocks:
            raise DocumentParseError(
                "NO_EXTRACTABLE_TEXT", "The document contains no extractable text."
            )
        return ParsedDocument(
            document_id=context.document_id,
            filename=context.filename,
            source_type=self.source_type,
            blocks=tuple(blocks),
            metadata={},
            warnings=tuple(warnings),
        )


def _preflight_worksheet_xml(payload: bytes) -> tuple[_WorksheetPreflight, ...]:
    """Bound physical OOXML work before openpyxl constructs any worksheet cells."""
    nodes = 0
    cells = 0
    total_merged_area = 0
    possible_materialized_cells = 0
    preflights: list[_WorksheetPreflight] = []
    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            for title, part_name in _worksheet_parts(archive):
                physical_coordinates: set[tuple[int, int]] = set()
                merged_ranges: set[tuple[int, int, int, int]] = set()
                materializing_ranges: set[tuple[int, int, int, int]] = set()
                with archive.open(part_name) as worksheet_xml:
                    for _event, element in etree.iterparse(
                        worksheet_xml,
                        events=("end",),
                        resolve_entities=False,
                        no_network=True,
                        huge_tree=False,
                    ):
                        nodes += 1
                        if nodes > settings.parser.max_spreadsheet_xml_nodes:
                            raise DocumentParseError(
                                "TABLE_XML_NODE_LIMIT_EXCEEDED",
                                "The spreadsheet exceeds the configured XML node limit.",
                            )
                        localname = etree.QName(element).localname
                        if localname == "c":
                            cells += 1
                            possible_materialized_cells += 1
                            _enforce_physical_cell_limit(cells)
                            _enforce_physical_cell_limit(possible_materialized_cells)
                            physical_coordinates.add(
                                _validate_physical_cell_coordinate(element)
                            )
                        elif localname == "mergeCell":
                            boundaries, area = _validated_range(element)
                            if area > settings.parser.max_merged_cell_area:
                                raise DocumentParseError(
                                    "TABLE_MERGED_CELL_LIMIT_EXCEEDED",
                                    "A merged spreadsheet range exceeds the configured area limit.",
                                )
                            total_merged_area += _unique_range_area(
                                boundaries, area, merged_ranges
                            )
                            if (
                                total_merged_area
                                > settings.parser.max_total_merged_cell_area
                            ):
                                raise DocumentParseError(
                                    "TABLE_TOTAL_MERGED_CELL_LIMIT_EXCEEDED",
                                    "The spreadsheet exceeds the configured merged-cell area limit.",
                                )
                            possible_materialized_cells += _unique_range_area(
                                boundaries, area, materializing_ranges
                            )
                            _enforce_physical_cell_limit(possible_materialized_cells)
                        elif localname == "hyperlink":
                            boundaries, area = _validated_range(element)
                            possible_materialized_cells += _unique_range_area(
                                boundaries, area, materializing_ranges
                            )
                            _enforce_physical_cell_limit(possible_materialized_cells)
                        if localname not in {"f", "v", "t", "is"}:
                            element.clear()
                preflights.append(
                    _WorksheetPreflight(
                        title=title,
                        part_name=part_name,
                        physical_coordinates=frozenset(physical_coordinates),
                    )
                )
    except DocumentParseError:
        raise
    except (BadZipFile, KeyError, OSError, ValueError, XMLSyntaxError) as error:
        raise _malformed_xlsx_error() from error
    return tuple(preflights)


def _worksheet_parts(archive: ZipFile) -> tuple[tuple[str, str], ...]:
    """Resolve workbook sheets to exact case-sensitive package members."""
    member_names = archive.namelist()
    if len(member_names) != len(set(member_names)):
        raise _malformed_xlsx_error()

    xml_parser = etree.XMLParser(
        resolve_entities=False,
        no_network=True,
        huge_tree=False,
    )
    workbook_root = etree.fromstring(archive.read(_WORKBOOK_PART), parser=xml_parser)
    relationships_root = etree.fromstring(
        archive.read(_WORKBOOK_RELATIONSHIPS_PART), parser=xml_parser
    )

    relationships: dict[str, Any] = {}
    for relationship in relationships_root:
        if etree.QName(relationship).localname != "Relationship":
            continue
        relationship_id = relationship.get("Id")
        if not relationship_id or relationship_id in relationships:
            raise _malformed_xlsx_error()
        relationships[relationship_id] = relationship

    resolved_parts: list[tuple[str, str]] = []
    seen_parts: set[str] = set()
    sheet_containers = [
        child
        for child in workbook_root
        if etree.QName(child).localname == "sheets"
    ]
    if len(sheet_containers) != 1:
        raise _malformed_xlsx_error()
    for sheet in sheet_containers[0]:
        if etree.QName(sheet).localname != "sheet":
            continue
        title = sheet.get("name")
        relationship_id = _sheet_relationship_id(sheet)
        if not title or not relationship_id or relationship_id not in relationships:
            raise _malformed_xlsx_error()
        relationship = relationships[relationship_id]
        if relationship.get("Type") not in _WORKSHEET_RELATIONSHIP_TYPES:
            raise _malformed_xlsx_error()
        if relationship.get("TargetMode") not in {None, "Internal"}:
            raise _malformed_xlsx_error()
        part_name = _resolve_worksheet_target(relationship.get("Target"), member_names)
        if part_name in seen_parts:
            raise _malformed_xlsx_error()
        seen_parts.add(part_name)
        resolved_parts.append((title, part_name))
    return tuple(resolved_parts)


def _sheet_relationship_id(sheet: Any) -> str | None:
    relationship_ids = [
        value
        for name, value in sheet.attrib.items()
        if etree.QName(name).localname == "id"
        and etree.QName(name).namespace in _RELATIONSHIP_NAMESPACES
    ]
    if len(relationship_ids) != 1:
        raise _malformed_xlsx_error()
    return relationship_ids[0]


def _resolve_worksheet_target(target: str | None, member_names: list[str]) -> str:
    if not target or "\\" in target or "%" in target:
        raise _malformed_xlsx_error()
    parsed = urlsplit(target)
    if parsed.scheme or parsed.netloc or parsed.query or parsed.fragment:
        raise _malformed_xlsx_error()
    package_path = target[1:] if target.startswith("/") else target
    segments = package_path.split("/")
    if any(segment in {"", ".", ".."} for segment in segments):
        raise _malformed_xlsx_error()
    if target.startswith("/"):
        resolved = PurePosixPath(*segments)
    else:
        resolved = PurePosixPath(_WORKBOOK_PART).parent.joinpath(*segments)
    part_name = resolved.as_posix()
    if part_name not in member_names:
        raise _malformed_xlsx_error()
    return part_name


def _validate_physical_cell_coordinate(element: Any) -> tuple[int, int]:
    coordinate = element.get("r")
    if not coordinate:
        raise _malformed_xlsx_error()
    try:
        row, column = coordinate_to_tuple(coordinate)
    except (TypeError, ValueError) as error:
        raise _malformed_xlsx_error() from error
    if not (1 <= row <= _XLSX_MAX_ROW and 1 <= column <= _XLSX_MAX_COLUMN):
        raise _malformed_xlsx_error()
    has_content = _physical_cell_has_content(element)
    if has_content and row > settings.parser.max_row_coordinate:
        raise DocumentParseError(
            "TABLE_ROW_LIMIT_EXCEEDED",
            "A non-empty spreadsheet cell exceeds the configured row coordinate limit.",
        )
    if has_content and column > min(
        settings.parser.max_columns, settings.parser.max_column_coordinate
    ):
        raise DocumentParseError(
            "TABLE_COLUMN_LIMIT_EXCEEDED",
            "A non-empty spreadsheet cell exceeds the configured column limit.",
        )
    return row, column


def _validated_range(element: Any) -> tuple[tuple[int, int, int, int], int]:
    reference = element.get("ref")
    if not reference:
        raise _malformed_xlsx_error()
    try:
        min_column, min_row, max_column, max_row = range_boundaries(reference)
    except (TypeError, ValueError) as error:
        raise _malformed_xlsx_error() from error
    boundaries = (min_column, min_row, max_column, max_row)
    if (
        not all(isinstance(value, int) for value in boundaries)
        or min_column < 1
        or min_row < 1
        or max_column < min_column
        or max_row < min_row
        or max_column > _XLSX_MAX_COLUMN
        or max_row > _XLSX_MAX_ROW
    ):
        raise _malformed_xlsx_error()
    area = (max_column - min_column + 1) * (max_row - min_row + 1)
    return boundaries, area


def _unique_range_area(
    boundaries: tuple[int, int, int, int],
    area: int,
    accounted_ranges: set[tuple[int, int, int, int]],
) -> int:
    if boundaries in accounted_ranges:
        return 0
    accounted_ranges.add(boundaries)
    return area


def _enforce_physical_cell_limit(possible_materialized_cells: int) -> None:
    if possible_materialized_cells > settings.parser.max_physical_cells:
        raise DocumentParseError(
            "TABLE_PHYSICAL_CELL_LIMIT_EXCEEDED",
            "The spreadsheet exceeds the configured physical cell limit.",
        )


def _physical_cell_has_content(element: Any) -> bool:
    for child in element.iterdescendants():
        localname = etree.QName(child).localname
        if localname == "f":
            return True
        if localname in {"v", "t"} and (child.text or "") != "":
            return True
    return False


def _worksheet_rows(
    worksheet: Any,
    physical_cells: Iterator[tuple[int, int, Any, Any]],
    warnings: list[ParserWarning],
) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """Yield sparse physical rows, never the rectangular declared dimension."""
    rows: dict[int, dict[int, Any]] = {}
    for source_row, source_column, cell, cached_cell in physical_cells:
        value = cell.value
        if cell.data_type == "f":
            cached = cached_cell.value if cached_cell is not None else None
            if cached is not None:
                value = cached
            else:
                value = value if str(value).startswith("=") else f"={value}"
                warnings.append(
                    ParserWarning(
                        "FORMULA_CACHE_UNAVAILABLE",
                        "A formula had no cached value; its formula text was retained.",
                        {"sheet": worksheet.title, "cell": cell.coordinate},
                    )
                )
        if value is not None:
            rows.setdefault(source_row, {})[source_column] = value
    for source_row in sorted(rows):
        columns = rows[source_row]
        last_column = max(columns)
        yield source_row, tuple(columns.get(column) for column in range(1, last_column + 1))


def _physical_cells(
    worksheet: Any,
    cached_worksheet: Any,
    physical_coordinates: frozenset[tuple[int, int]],
) -> Iterator[tuple[int, int, Any, Any]]:
    """Adapt bounded OOXML coordinates to OpenPyXL's isolated private mappings."""
    formula_cells = getattr(worksheet, "_cells", None)
    cached_cells = getattr(cached_worksheet, "_cells", None)
    if not isinstance(formula_cells, Mapping) or not isinstance(cached_cells, Mapping):
        raise _malformed_xlsx_error()
    for cell_mapping in (formula_cells, cached_cells):
        for coordinate in cell_mapping:
            if not _valid_mapping_coordinate(coordinate):
                raise _malformed_xlsx_error()
            if coordinate not in physical_coordinates:
                unexpected_cell = cell_mapping.get(coordinate)
                if not (
                    isinstance(unexpected_cell, MergedCell)
                    or (
                        isinstance(unexpected_cell, Cell)
                        and unexpected_cell.hyperlink is not None
                    )
                ):
                    raise _malformed_xlsx_error()
    for row, column in sorted(physical_coordinates):
        cell = formula_cells.get((row, column))
        cached_cell = cached_cells.get((row, column))
        if isinstance(cell, MergedCell):
            continue
        if cell is None:
            continue
        if not isinstance(cell, Cell):
            raise _malformed_xlsx_error()
        if isinstance(cached_cell, MergedCell):
            cached_cell = None
        elif cached_cell is not None and not isinstance(cached_cell, Cell):
            raise _malformed_xlsx_error()
        yield row, column, cell, cached_cell


def _valid_mapping_coordinate(coordinate: Any) -> bool:
    return (
        isinstance(coordinate, tuple)
        and len(coordinate) == 2
        and all(isinstance(value, int) for value in coordinate)
        and 1 <= coordinate[0] <= _XLSX_MAX_ROW
        and 1 <= coordinate[1] <= _XLSX_MAX_COLUMN
    )


def _hidden_sheet_warning(sheet: str) -> ParserWarning:
    return ParserWarning(
        "HIDDEN_SHEET_SKIPPED",
        "A hidden worksheet was skipped.",
        {"sheet": sheet},
    )


def _empty_sheet_warning(sheet: str) -> ParserWarning:
    return ParserWarning(
        "EMPTY_SHEET_SKIPPED",
        "A worksheet without data rows was skipped.",
        {"sheet": sheet},
    )


def _malformed_xlsx_error() -> DocumentParseError:
    return DocumentParseError("XLSX_MALFORMED", "The XLSX file is malformed or unreadable.")
