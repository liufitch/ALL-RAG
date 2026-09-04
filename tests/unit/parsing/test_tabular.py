import io
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
import xlwt
from lxml import etree
from openpyxl import Workbook

from rag_modules.config.settings import settings
from rag_modules.parsing.base import ParseContext
from rag_modules.parsing.csv_parser import CsvParser
from rag_modules.parsing.models import DocumentParseError
from rag_modules.parsing.registry import ParserRegistry
from rag_modules.parsing.xls_parser import XlsParser
from rag_modules.parsing.xlsx_parser import XlsxParser


FIXTURE_DIRECTORY = Path(__file__).parents[2] / "fixtures" / "documents"


@pytest.fixture
def fixture_bytes():
    def load(name: str) -> bytes:
        return (FIXTURE_DIRECTORY / name).read_bytes()

    return load


@pytest.fixture
def parser_registry() -> ParserRegistry:
    return ParserRegistry(
        {".xls": XlsParser(), ".xlsx": XlsxParser(), ".csv": CsvParser()}
    )


def _xlsx_bytes(
    configure,
    *,
    sheet_xml=None,
    worksheet_member: str | None = None,
    workbook_rels_xml=None,
) -> bytes:
    output = io.BytesIO()
    workbook = Workbook()
    configure(workbook)
    workbook.save(output)
    if sheet_xml is None and worksheet_member is None and workbook_rels_xml is None:
        return output.getvalue()
    patched = io.BytesIO()
    with ZipFile(io.BytesIO(output.getvalue())) as source, ZipFile(
        patched, "w", ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                if sheet_xml is not None:
                    payload = sheet_xml(payload)
                if worksheet_member is not None:
                    target.writestr(worksheet_member, payload)
                    continue
            if info.filename == "xl/_rels/workbook.xml.rels":
                if worksheet_member is not None:
                    package_target = f"/{worksheet_member}".encode()
                    relative_target = worksheet_member.removeprefix("xl/").encode()
                    payload = payload.replace(
                        b"/xl/worksheets/sheet1.xml", package_target
                    ).replace(b"worksheets/sheet1.xml", relative_target)
                if workbook_rels_xml is not None:
                    payload = workbook_rels_xml(payload)
            target.writestr(info, payload)
    return patched.getvalue()


def _modify_first_worksheet_relationship(payload: bytes, **attributes) -> bytes:
    root = etree.fromstring(payload)
    relationship = next(
        child
        for child in root
        if child.get("Type", "").endswith("/relationships/worksheet")
    )
    for name, value in attributes.items():
        if value is None:
            relationship.attrib.pop(name, None)
        else:
            relationship.set(name, value)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def _duplicate_worksheet_target(payload: bytes) -> bytes:
    root = etree.fromstring(payload)
    worksheet_relationships = [
        child
        for child in root
        if child.get("Type", "").endswith("/relationships/worksheet")
    ]
    worksheet_relationships[1].set("Target", worksheet_relationships[0].get("Target"))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def _duplicate_relationship_id(payload: bytes) -> bytes:
    root = etree.fromstring(payload)
    root[1].set("Id", root[0].get("Id"))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def _replace_merge_reference(payload: bytes, reference: str | None) -> bytes:
    root = etree.fromstring(payload)
    merge_cell = next(
        element
        for element in root.iter()
        if etree.QName(element).localname == "mergeCell"
    )
    if reference is None:
        merge_cell.attrib.pop("ref", None)
    else:
        merge_cell.set("ref", reference)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def _append_hyperlink(payload: bytes, reference: str | None) -> bytes:
    root = etree.fromstring(payload)
    namespace = etree.QName(root).namespace
    hyperlinks = etree.SubElement(root, f"{{{namespace}}}hyperlinks")
    attributes = {"location": "Sheet!A1"}
    if reference is not None:
        attributes["ref"] = reference
    etree.SubElement(hyperlinks, f"{{{namespace}}}hyperlink", **attributes)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def _xls_bytes(configure) -> bytes:
    output = io.BytesIO()
    workbook = xlwt.Workbook()
    configure(workbook)
    workbook.save(output)
    return output.getvalue()


@pytest.mark.parametrize("fixture_name", ["orders.xls", "orders.xlsx", "orders.csv"])
def test_tabular_parsers_emit_self_describing_rows(
    fixture_name, parser_registry, fixture_bytes
):
    """Dropping headers or source rows makes table facts ambiguous during indexing."""
    parsed = parser_registry.parse(
        Path(fixture_name).suffix,
        io.BytesIO(fixture_bytes(fixture_name)),
        ParseContext("d", fixture_name),
    )

    row = next(block for block in parsed.blocks if block.block_type == "table_row")
    assert row.text == "订单号：A001；客户：张三；金额：100"
    assert row.metadata == {
        "sheet": "CSV" if fixture_name.endswith(".csv") else "订单",
        "row": 2,
        "headers": ["订单号", "客户", "金额"],
    }


def test_xlsx_uses_sheet_name_skips_hidden_sheets_and_ignores_empty_sheets(
    parser_registry, fixture_bytes
):
    """Reading hidden or empty sheets would expose non-user table content as facts."""
    parsed = parser_registry.parse(
        ".xlsx",
        io.BytesIO(fixture_bytes("multi-sheet.xlsx")),
        ParseContext("d", "x.xlsx"),
    )

    assert {block.metadata["sheet"] for block in parsed.blocks} == {"订单", "客户"}
    assert any(warning.code == "HIDDEN_SHEET_SKIPPED" for warning in parsed.warnings)
    assert all(block.metadata["sheet"] != "空白" for block in parsed.blocks)


def test_tabular_normalization_keeps_column_alignment_and_formula_text(
    parser_registry, fixture_bytes
):
    """Coercing nulls or cached formula values would pair facts with the wrong headers."""
    parsed = parser_registry.parse(
        ".xlsx",
        io.BytesIO(fixture_bytes("normalization.xlsx")),
        ParseContext("d", "normalization.xlsx"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        (
            "日期：2026-08-31；开关：true；金额：100；列4：保留；名称_2：=SUM(1,2)",
            {
                "sheet": "数据",
                "row": 2,
                "headers": ["日期", "开关", "金额", "列4", "名称", "名称_2"],
            },
        )
    ]


def test_csv_uses_sniffed_semicolon_dialect_and_true_line_number_after_quoted_newline(
    parser_registry, fixture_bytes
):
    """Using record indexes would mislocate CSV data following a quoted newline."""
    parsed = parser_registry.parse(
        ".csv",
        io.BytesIO(fixture_bytes("quoted-newline.csv")),
        ParseContext("d", "quoted-newline.csv"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        (
            "名称：第一\n行；备注：ok",
            {"sheet": "CSV", "row": 3, "headers": ["名称", "备注"]},
        ),
        (
            "名称：第二行；备注：done",
            {"sheet": "CSV", "row": 4, "headers": ["名称", "备注"]},
        ),
    ]


def test_csv_falls_back_deterministically_for_a_short_single_column_file(
    parser_registry, fixture_bytes
):
    """Letting Sniffer failure escape would reject valid one-column CSV uploads."""
    parsed = parser_registry.parse(
        ".csv",
        io.BytesIO(fixture_bytes("short-single-column.csv")),
        ParseContext("d", "single.csv"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        ("name：alice", {"sheet": "CSV", "row": 2, "headers": ["name"]})
    ]


def test_empty_xlsx_has_no_extractable_table_rows(parser_registry, fixture_bytes):
    """Treating an empty sheet as a data row would fabricate searchable table content."""
    with pytest.raises(DocumentParseError) as error:
        parser_registry.parse(
            ".xlsx",
            io.BytesIO(fixture_bytes("empty.xlsx")),
            ParseContext("d", "empty.xlsx"),
        )

    assert error.value.code == "NO_EXTRACTABLE_TEXT"


def test_xls_normalizes_blank_cells_and_booleans_without_emitting_empty_rows(
    parser_registry, fixture_bytes
):
    """Passing xlrd blank/boolean primitives through creates false table facts."""
    parsed = parser_registry.parse(
        ".xls",
        io.BytesIO(fixture_bytes("legacy-null-bool.xls")),
        ParseContext("d", "legacy-null-bool.xls"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        (
            "左：A；右：C；启用：true",
            {"sheet": "旧表", "row": 2, "headers": ["左", "中", "右", "启用"]},
        )
    ]


def test_csv_normalizes_empty_fields_and_ignores_a_fully_empty_data_row(
    parser_registry, fixture_bytes
):
    """Treating CSV empty fields as text emits empty pairs and a fake data record."""
    parsed = parser_registry.parse(
        ".csv",
        io.BytesIO(fixture_bytes("null-rows.csv")),
        ParseContext("d", "null-rows.csv"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        (
            "左：A；右：C",
            {"sheet": "CSV", "row": 2, "headers": ["左", "中", "右"]},
        )
    ]


def test_ragged_csv_extends_and_deduplicates_headers_before_any_block_metadata(
    parser_registry, fixture_bytes
):
    """Finalizing ragged rows early leaves metadata unable to describe emitted columns."""
    parsed = parser_registry.parse(
        ".csv",
        io.BytesIO(fixture_bytes("ragged.csv")),
        ParseContext("d", "ragged.csv"),
    )

    assert [(block.text, block.metadata) for block in parsed.blocks] == [
        (
            "甲：A；列3：B；列3_2：C",
            {"sheet": "CSV", "row": 2, "headers": ["甲", "列3", "列3_2"]},
        )
    ]


def test_row_limit_is_a_document_budget_that_counts_visible_sheet_headers(
    monkeypatch, parser_registry, fixture_bytes
):
    """Resetting the limit per sheet permits a workbook to exceed its document budget."""
    monkeypatch.setattr(settings.parser, "max_rows", 3)

    with pytest.raises(DocumentParseError) as error:
        parser_registry.parse(
            ".xlsx",
            io.BytesIO(fixture_bytes("row-budget.xlsx")),
            ParseContext("d", "row-budget.xlsx"),
        )

    assert error.value.code == "TABLE_ROW_LIMIT_EXCEEDED"


def test_csv_row_budget_counts_non_empty_header_and_data_but_not_blank_records(
    monkeypatch, parser_registry, fixture_bytes
):
    """A blank CSV record must not consume the same logical row budget as source data."""
    monkeypatch.setattr(settings.parser, "max_rows", 2)

    parsed = parser_registry.parse(
        ".csv",
        io.BytesIO(fixture_bytes("null-rows.csv")),
        ParseContext("d", "null-rows.csv"),
    )

    assert [block.metadata["row"] for block in parsed.blocks] == [2]


def test_xlsx_ignores_style_only_declared_dimensions_for_row_and_column_limits(
    monkeypatch, parser_registry, fixture_bytes
):
    """Trusting a worksheet's declared Z100 dimension rejects tiny real tables."""
    monkeypatch.setattr(settings.parser, "max_rows", 2)
    monkeypatch.setattr(settings.parser, "max_columns", 2)

    parsed = parser_registry.parse(
        ".xlsx",
        io.BytesIO(fixture_bytes("style-only-dimensions.xlsx")),
        ParseContext("d", "style-only-dimensions.xlsx"),
    )

    assert [(block.text, block.metadata["row"]) for block in parsed.blocks] == [("列：值", 2)]


def test_xlsx_remote_non_null_cell_uses_its_real_column_for_the_limit(
    monkeypatch, parser_registry, fixture_bytes
):
    """Ignoring declared dimensions must not hide a real distant value beyond the column limit."""
    monkeypatch.setattr(settings.parser, "max_rows", 2)
    monkeypatch.setattr(settings.parser, "max_columns", 2)

    with pytest.raises(DocumentParseError) as error:
        parser_registry.parse(
            ".xlsx",
            io.BytesIO(fixture_bytes("remote-value-dimensions.xlsx")),
            ParseContext("d", "remote-value-dimensions.xlsx"),
        )

    assert error.value.code == "TABLE_COLUMN_LIMIT_EXCEEDED"


def test_xlsx_forged_declared_dimension_never_expands_empty_coordinate_space(
    monkeypatch,
):
    """A forged A1:XFD1048576 dimension must not drive dense row iteration."""
    monkeypatch.setattr(settings.parser, "max_rows", 2)
    monkeypatch.setattr(settings.parser, "max_columns", 2)

    payload = _xlsx_bytes(
        lambda workbook: (
            setattr(workbook.active, "title", "Data"),
            workbook.active.append(["列"]),
            workbook.active.append(["值"]),
        ),
        sheet_xml=lambda xml: xml.replace(
            b'<dimension ref="A1:A2"/>', b'<dimension ref="A1:XFD1048576"/>'
        ),
    )
    parsed = XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "forged.xlsx"))

    assert [block.text for block in parsed.blocks] == ["列：值"]


def test_xlsx_rejects_remote_nonempty_row_without_dense_iteration(monkeypatch):
    """A real distant cell must fail stably without allocating all intervening rows."""
    monkeypatch.setattr(settings.parser, "max_row_coordinate", 100)
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["列"]),
            workbook.active.__setitem__("A1048576", "远端"),
        )
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "remote.xlsx"))

    assert error.value.code == "TABLE_ROW_LIMIT_EXCEEDED"


def test_xlsx_enforces_independent_physical_cell_budget(monkeypatch):
    """Physical OOXML cells are bounded independently from meaningful table rows."""
    monkeypatch.setattr(settings.parser, "max_physical_cells", 2)
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["列1", "列2"]),
            workbook.active.append(["值", None]),
        )
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "cells.xlsx"))

    assert error.value.code == "TABLE_PHYSICAL_CELL_LIMIT_EXCEEDED"


def test_xlsx_case_varied_relationship_target_is_still_preflighted(monkeypatch):
    """Discovering worksheets by a lowercase filename glob skips referenced parts."""
    monkeypatch.setattr(settings.parser, "max_physical_cells", 2)
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["列1", "列2"]),
            workbook.active.append(["值"]),
        ),
        worksheet_member="xl/worksheets/sheet1.XML",
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran before relationship-resolved preflight")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "case-varied.xlsx"))

    assert error.value.code == "TABLE_PHYSICAL_CELL_LIMIT_EXCEEDED"


@pytest.mark.parametrize(
    "workbook_rels_xml",
    [
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, TargetMode="External"
            ),
            id="external-relationship",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="//fileserver/share/sheet1.xml"
            ),
            id="absolute-authority-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="file:///tmp/sheet1.xml"
            ),
            id="uri-scheme-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="C:/tmp/sheet1.xml"
            ),
            id="drive-path-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target=r"worksheets\sheet1.xml"
            ),
            id="backslash-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets/sheet1.xml?download=1"
            ),
            id="query-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets/sheet1.xml#fragment"
            ),
            id="fragment-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets/%73heet1.xml"
            ),
            id="percent-escape-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets/./sheet1.xml"
            ),
            id="single-dot-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets//sheet1.xml"
            ),
            id="empty-segment-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="../worksheets/sheet1.xml"
            ),
            id="traversal-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload, Target="worksheets/missing.xml"
            ),
            id="missing-target",
        ),
        pytest.param(
            lambda payload: _modify_first_worksheet_relationship(
                payload,
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles",
            ),
            id="non-worksheet-relationship",
        ),
        pytest.param(_duplicate_relationship_id, id="duplicate-relationship-id"),
    ],
)
def test_xlsx_relationship_corruption_is_rejected_before_openpyxl(
    monkeypatch, workbook_rels_xml
):
    """Untrusted workbook relationships must not reach OpenPyXL construction."""
    payload = _xlsx_bytes(
        lambda workbook: workbook.active.append(["列"]),
        workbook_rels_xml=workbook_rels_xml,
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran for a malformed relationship package")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "relationships.xlsx"))

    assert error.value.code == "XLSX_MALFORMED"
    assert error.value.message == "The XLSX file is malformed or unreadable."


def test_xlsx_accepts_one_leading_slash_as_a_package_root_target():
    """Treating an OPC package-root target as a host path rejects normal XLSX files."""
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["列"]),
            workbook.active.append(["值"]),
        ),
        workbook_rels_xml=lambda xml: _modify_first_worksheet_relationship(
            xml, Target="/xl/worksheets/sheet1.xml"
        ),
    )

    parsed = XlsxParser().parse(
        io.BytesIO(payload), ParseContext("d", "package-root.xlsx")
    )

    assert [block.text for block in parsed.blocks] == ["列：值"]


def test_xlsx_duplicate_worksheet_target_is_rejected_before_openpyxl(monkeypatch):
    """Two workbook sheets resolving to one part make sheet identity ambiguous."""
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["第一"]),
            workbook.create_sheet("Second").append(["第二"]),
        ),
        workbook_rels_xml=_duplicate_worksheet_target,
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran for duplicate worksheet targets")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "duplicate-target.xlsx"))

    assert error.value.code == "XLSX_MALFORMED"


def test_xlsx_rejects_single_merged_range_area_before_openpyxl(monkeypatch):
    """One sparse merge declaration must not expand into unbounded placeholders."""
    monkeypatch.setattr(settings.parser, "max_merged_cell_area", 100)
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.__setitem__("A1", "列"),
            workbook.active.merge_cells("A1:B2"),
        ),
        sheet_xml=lambda xml: _replace_merge_reference(xml, "A1:Z100"),
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran before the single-merge bound")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "merge-area.xlsx"))

    assert error.value.code == "TABLE_MERGED_CELL_LIMIT_EXCEEDED"
    assert (
        error.value.message
        == "A merged spreadsheet range exceeds the configured area limit."
    )


def test_xlsx_rejects_total_merged_range_area_before_openpyxl(monkeypatch):
    """Individually legal ranges must not bypass the document-wide merge bound."""
    monkeypatch.setattr(settings.parser, "max_merged_cell_area", 4)
    monkeypatch.setattr(settings.parser, "max_total_merged_cell_area", 10)

    def configure(workbook):
        worksheet = workbook.active
        for coordinate in ("A1", "C1", "E1"):
            worksheet[coordinate] = coordinate
        for reference in ("A1:B2", "C1:D2", "E1:F2"):
            worksheet.merge_cells(reference)

    payload = _xlsx_bytes(configure)

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran before the aggregate-merge bound")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "merge-total.xlsx"))

    assert error.value.code == "TABLE_TOTAL_MERGED_CELL_LIMIT_EXCEEDED"
    assert (
        error.value.message
        == "The spreadsheet exceeds the configured merged-cell area limit."
    )


@pytest.mark.parametrize(
    "reference",
    [
        pytest.param("not-a-range", id="malformed-merge-reference"),
        pytest.param("B2:A1", id="reversed-merge-reference"),
        pytest.param("XFE1:XFE1", id="out-of-sheet-merge-reference"),
        pytest.param(None, id="missing-merge-reference"),
    ],
)
def test_xlsx_rejects_malformed_merged_range_before_openpyxl(
    monkeypatch, reference
):
    """Invalid merge coordinates must fail at the bounded XML boundary."""
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.__setitem__("A1", "列"),
            workbook.active.merge_cells("A1:B2"),
        ),
        sheet_xml=lambda xml: _replace_merge_reference(xml, reference),
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran for an invalid merge range")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "bad-merge.xlsx"))

    assert error.value.code == "XLSX_MALFORMED"
    assert error.value.message == "The XLSX file is malformed or unreadable."


def test_xlsx_physical_cell_adapter_skips_merged_placeholders_without_growth():
    """Walking preflight coordinates must not emit or create merge placeholders."""
    from openpyxl import load_workbook as openpyxl_load_workbook

    from rag_modules.parsing.xlsx_parser import (
        _physical_cells,
        _preflight_worksheet_xml,
    )

    def configure(workbook):
        worksheet = workbook.active
        worksheet["A1"] = "分组"
        worksheet["C1"] = "类别"
        worksheet.merge_cells("A1:B2")
        worksheet["A3"] = "值"
        worksheet["C3"] = "甲"

    payload = _xlsx_bytes(configure)
    preflight = _preflight_worksheet_xml(payload)[0]
    formula_workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=False)
    cached_workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    formula_sheet = formula_workbook.active
    cached_sheet = cached_workbook.active
    before_sizes = (len(formula_sheet._cells), len(cached_sheet._cells))
    try:
        physical_cells = list(
            _physical_cells(
                formula_sheet,
                cached_sheet,
                preflight.physical_coordinates,
            )
        )
    finally:
        formula_workbook.close()
        cached_workbook.close()

    assert [(row, column) for row, column, _cell, _cached in physical_cells] == [
        (1, 1),
        (1, 3),
        (3, 1),
        (3, 3),
    ]
    assert all(type(cell).__name__ == "Cell" for _, _, cell, _ in physical_cells)
    assert (len(formula_sheet._cells), len(cached_sheet._cells)) == before_sizes


def test_xlsx_normal_merged_range_emits_only_physical_values():
    """A legal merge remains parseable without placeholder rows or values."""
    def configure(workbook):
        worksheet = workbook.active
        worksheet["A1"] = "分组"
        worksheet["C1"] = "类别"
        worksheet.merge_cells("A1:B2")
        worksheet["A3"] = "值"
        worksheet["C3"] = "甲"

    parsed = XlsxParser().parse(
        io.BytesIO(_xlsx_bytes(configure)), ParseContext("d", "merged.xlsx")
    )

    assert [block.text for block in parsed.blocks] == ["分组：值；类别：甲"]
    assert [block.metadata["row"] for block in parsed.blocks] == [3]


def test_xlsx_rejects_hyperlink_range_materialization_before_openpyxl(monkeypatch):
    """A hyperlink range must share the physical-cell materialization budget."""
    monkeypatch.setattr(settings.parser, "max_physical_cells", 4)
    payload = _xlsx_bytes(
        lambda workbook: workbook.active.__setitem__("A1", "列"),
        sheet_xml=lambda xml: _append_hyperlink(xml, "B2:C3"),
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran before the hyperlink materialization bound")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "hyperlink-limit.xlsx"))

    assert error.value.code == "TABLE_PHYSICAL_CELL_LIMIT_EXCEEDED"
    assert (
        error.value.message
        == "The spreadsheet exceeds the configured physical cell limit."
    )


@pytest.mark.parametrize(
    "reference",
    [
        pytest.param("not-a-range", id="malformed-hyperlink-reference"),
        pytest.param("C3:B2", id="reversed-hyperlink-reference"),
        pytest.param("XFE1:XFE1", id="out-of-sheet-hyperlink-reference"),
        pytest.param(None, id="missing-hyperlink-reference"),
    ],
)
def test_xlsx_rejects_malformed_hyperlink_before_openpyxl(
    monkeypatch, reference
):
    """Invalid hyperlink ranges must fail at the bounded XML boundary."""
    payload = _xlsx_bytes(
        lambda workbook: workbook.active.__setitem__("A1", "列"),
        sheet_xml=lambda xml: _append_hyperlink(xml, reference),
    )

    def fail_if_openpyxl_runs(*args, **kwargs):
        pytest.fail("load_workbook ran for an invalid hyperlink range")

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook", fail_if_openpyxl_runs
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "bad-hyperlink.xlsx"))

    assert error.value.code == "XLSX_MALFORMED"
    assert error.value.message == "The XLSX file is malformed or unreadable."


def test_xlsx_bounded_hyperlink_range_does_not_become_table_data(monkeypatch):
    """Bounded hyperlink-created cells are valid but are not physical table values."""
    monkeypatch.setattr(settings.parser, "max_physical_cells", 6)
    payload = _xlsx_bytes(
        lambda workbook: (
            workbook.active.append(["列"]),
            workbook.active.append(["值"]),
        ),
        sheet_xml=lambda xml: _append_hyperlink(xml, "B2:C3"),
    )

    parsed = XlsxParser().parse(
        io.BytesIO(payload), ParseContext("d", "bounded-hyperlink.xlsx")
    )

    assert [block.text for block in parsed.blocks] == ["列：值"]


def test_xlsx_hidden_sheet_mapping_contract_is_validated(monkeypatch):
    """Skipping hidden extraction must not skip private-mapping validation."""
    from openpyxl import load_workbook as openpyxl_load_workbook

    def configure(workbook):
        workbook.active.title = "Data"
        workbook.active.append(["列"])
        workbook.active.append(["值"])
        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "secret"

    payload = _xlsx_bytes(configure)

    def load_with_corrupt_hidden_mapping(*args, **kwargs):
        workbook = openpyxl_load_workbook(*args, **kwargs)
        workbook["Hidden"]._cells[(5, 5)] = object()
        return workbook

    monkeypatch.setattr(
        "rag_modules.parsing.xlsx_parser.load_workbook",
        load_with_corrupt_hidden_mapping,
    )

    with pytest.raises(DocumentParseError) as error:
        XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "hidden-contract.xlsx"))

    assert error.value.code == "XLSX_MALFORMED"


def test_xls_and_xlsx_warn_for_hidden_very_hidden_and_nondata_sheets():
    """Hidden sheets and visible sheets without data rows must be skipped with warnings."""
    def xlsx_workbook(workbook):
        visible = workbook.active
        visible.title = "Data"
        visible.append(["列"])
        visible.append(["值"])
        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden.append(["secret"])
        very_hidden = workbook.create_sheet("VeryHidden")
        very_hidden.sheet_state = "veryHidden"
        very_hidden.append(["secret"])
        workbook.create_sheet("Empty")
        header = workbook.create_sheet("HeaderOnly")
        header.append(["列"])

    def xls_workbook(workbook):
        data = workbook.add_sheet("Data")
        data.write(0, 0, "列")
        data.write(1, 0, "值")
        hidden = workbook.add_sheet("Hidden")
        hidden.visibility = 1
        hidden.write(0, 0, "secret")
        very_hidden = workbook.add_sheet("VeryHidden")
        very_hidden.visibility = 2
        very_hidden.write(0, 0, "secret")
        workbook.add_sheet("Empty")
        header = workbook.add_sheet("HeaderOnly")
        header.write(0, 0, "列")

    for parser, payload, filename in (
        (XlsxParser(), _xlsx_bytes(xlsx_workbook), "sheets.xlsx"),
        (XlsParser(), _xls_bytes(xls_workbook), "sheets.xls"),
    ):
        parsed = parser.parse(io.BytesIO(payload), ParseContext("d", filename))
        assert [block.text for block in parsed.blocks] == ["列：值"]
        assert [warning.code for warning in parsed.warnings].count("HIDDEN_SHEET_SKIPPED") == 2
        assert [warning.code for warning in parsed.warnings].count("EMPTY_SHEET_SKIPPED") == 2


def test_xlsx_prefers_real_cached_formula_and_warns_when_cache_is_absent():
    """Formula text is only a fallback; a genuine OOXML cached value wins."""
    def workbook_with_formulas(workbook):
        workbook.active.append(["缓存", "无缓存"])
        workbook.active.append(["=1+2", "=2+3"])

    payload = _xlsx_bytes(
        workbook_with_formulas,
        sheet_xml=lambda xml: xml.replace(
            b"<f>1+2</f><v></v>", b"<f>1+2</f><v>3</v>"
        ),
    )
    with ZipFile(io.BytesIO(payload)) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml")
    assert b"<f>1+2</f><v>3</v>" in xml
    assert b"<f>2+3</f><v></v>" in xml

    parsed = XlsxParser().parse(io.BytesIO(payload), ParseContext("d", "formula.xlsx"))

    assert [block.text for block in parsed.blocks] == ["缓存：3；无缓存：=2+3"]
    assert [warning.code for warning in parsed.warnings] == ["FORMULA_CACHE_UNAVAILABLE"]


def test_xlsx_aggregates_formula_warnings_per_sheet_with_bounded_coordinate_samples(
    monkeypatch,
):
    """One missing-cache warning per cell would retain formula-scale metadata."""
    monkeypatch.setattr(settings.parser, "max_formula_warning_samples", 5)

    def workbook_with_many_formulas(workbook):
        sheet = workbook.active
        sheet.title = "Calculations"
        sheet.append(["结果"])
        for index in range(1_000):
            sheet.append([f"={index}+1"])

    parsed = XlsxParser().parse(
        io.BytesIO(_xlsx_bytes(workbook_with_many_formulas)),
        ParseContext("d", "many-formulas.xlsx"),
    )

    assert len(parsed.warnings) == 1
    warning = parsed.warnings[0]
    assert warning.code == "FORMULA_CACHE_UNAVAILABLE"
    assert warning.metadata == {
        "sheet": "Calculations",
        "count": 1_000,
        "sample_cells": ["A2", "A3", "A4", "A5", "A6"],
    }
    assert "=" not in repr(warning.metadata)
    assert len(parsed.warnings) <= settings.parser.max_warnings_per_document


def test_xlsx_formula_warnings_are_per_sheet_and_share_the_document_warning_cap(
    monkeypatch,
):
    """Sheet aggregation must still leave one deterministic document-level bound."""
    monkeypatch.setattr(settings.parser, "max_warnings_per_document", 3)

    def workbook_with_two_formula_sheets(workbook):
        first = workbook.active
        first.title = "First"
        first.append(["结果"])
        first.append(["=1+1"])
        first.append(["=2+2"])
        second = workbook.create_sheet("Second")
        second.append(["=3+3"])

    parsed = XlsxParser().parse(
        io.BytesIO(_xlsx_bytes(workbook_with_two_formula_sheets)),
        ParseContext("d", "two-formula-sheets.xlsx"),
    )

    assert [warning.code for warning in parsed.warnings] == [
        "FORMULA_CACHE_UNAVAILABLE",
        "FORMULA_CACHE_UNAVAILABLE",
        "WARNINGS_TRUNCATED",
    ]
    assert [warning.metadata["sheet"] for warning in parsed.warnings[:2]] == [
        "First",
        "Second",
    ]
    assert [warning.metadata["count"] for warning in parsed.warnings[:2]] == [2, 1]
    assert parsed.warnings[-1].metadata == {"omitted_count": 1}


def test_xlsx_hidden_formula_sheet_has_only_the_hidden_sheet_warning():
    """Hidden content is validated but not extracted or reported as visible formulas."""
    def workbook_with_hidden_formula(workbook):
        visible = workbook.active
        visible.title = "Visible"
        visible.append(["列"])
        visible.append(["值"])
        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden.append(["=1+1"])

    parsed = XlsxParser().parse(
        io.BytesIO(_xlsx_bytes(workbook_with_hidden_formula)),
        ParseContext("d", "hidden-formula.xlsx"),
    )

    assert [warning.code for warning in parsed.warnings] == ["HIDDEN_SHEET_SKIPPED"]


@pytest.mark.parametrize(
    ("fixture_name", "limit_name", "limit", "expected_code"),
    [
        ("orders.xlsx", "max_rows", 1, "TABLE_ROW_LIMIT_EXCEEDED"),
        ("orders.xlsx", "max_columns", 2, "TABLE_COLUMN_LIMIT_EXCEEDED"),
        ("orders.csv", "max_cell_characters", 2, "TABLE_CELL_CHARACTER_LIMIT_EXCEEDED"),
    ],
)
def test_tabular_parsers_reject_configured_limits_at_the_parser_boundary(
    monkeypatch, parser_registry, fixture_bytes, fixture_name, limit_name, limit, expected_code
):
    """Silently truncating table bounds would let callers index incomplete source data."""
    monkeypatch.setattr(settings.parser, limit_name, limit)

    with pytest.raises(DocumentParseError) as error:
        parser_registry.parse(
            Path(fixture_name).suffix,
            io.BytesIO(fixture_bytes(fixture_name)),
            ParseContext("d", fixture_name),
        )

    assert error.value.code == expected_code
    assert error.value.retryable is False


@pytest.mark.parametrize(
    ("extension", "payload", "expected_code"),
    [
        (".xlsx", b"not a workbook", "XLSX_MALFORMED"),
        (".xls", b"not a workbook", "XLS_MALFORMED"),
        (".csv", bytes(range(256)) * 4, "CSV_ENCODING_UNCERTAIN"),
    ],
)
def test_tabular_parsers_normalize_unreadable_inputs_to_stable_errors(
    parser_registry, extension, payload, expected_code
):
    """Leaking reader-library exceptions would make upload failures API-unstable."""
    with pytest.raises(DocumentParseError) as error:
        parser_registry.parse(extension, io.BytesIO(payload), ParseContext("d", f"x{extension}"))

    assert error.value.code == expected_code
